from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
from cousine.models import Category, Ingredient, Nutritionals
from supply.models import Supplier, Article


def import_suppliers():
    wb=load_workbook('/home/mate/Documenti/supply_supplier.xlsx')
    ws=wb.active
    print(ws)
    for row in range(2,13):
        name=ws["B"+str(row)].value

        supplier=Supplier()
        supplier.name=name
        print(supplier)
        supplier.save()



def import_categories():
    wb=load_workbook('/home/mate/Documenti/food_category.xlsx')
    ws=wb.active
    for row in range(2,22):
            name=ws["B"+str(row)].value
            category=Category()
            category.name=name
            category.save()

    categories=Category.objects.all()
    print(categories)

def import_ingredients():
    wb=load_workbook('/home/mate/Documenti/food_ingredient.xlsx')
    ws=wb.active
    for row in range(2,160):
        name=ws["B"+str(row)].value
        ingredient=Ingredient()
        ingredient.name=name
        cat_id=int(ws["C"+str(row)].value)
        category=Category.objects.get(id=cat_id)
        ingredient.category=category
        nutritionals= Nutritionals()
        nutritionals.fat=0
        nutritionals.calories=0
        nutritionals.proteins=0
        nutritionals.carbohydrates=0
        nutritionals.starch=0
        nutritionals.save()
        ingredient.nutritional=nutritionals
        print(ingredient.name+ ingredient.category.name)
        ingredient.save()


def import_articles():
    wb=load_workbook('/home/mate/Documenti/supply_article.xlsx')
    ws=wb.active
    print(ws)
    for row in range(2,190):
        try:
            name=ws["B"+str(row)].value
            ing_id=ws["C"+str(row)].value
            ingredient=Ingredient.objects.get(pk=int(ing_id))
            ua=ws["D"+str(row)].value
            supp_id=ws["E"+str(row)].value
            supplier=Supplier.objects.get(id=supp_id)
            giac=ws["F"+str(row)].value
            price_alKg=ws["G"+str(row)].value
            article=Article()
            article.name=name
            article.ingredient=ingredient
            article.supplier=supplier
            article.unitaArrivo=ua
            article.giacenza_in_gr=giac
            article.prezzo_al_kg=price_alKg
            print(article.name)

            #article.save()
        except:
            print('e'+str(ws["A"+str(row)].value))
